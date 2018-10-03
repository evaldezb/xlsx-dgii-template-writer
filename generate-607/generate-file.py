
import os
import shutil
from openpyxl import Workbook
from openpyxl import load_workbook
from decorators import timer


destination_file = 'Preparació Envio Formato 607 - 2.xlsx'
file_name = 'Preparació Envio Formato 607.xlsx'


@timer
def get_template(file_name):
    new_file = Workbook(write_only=True)
    print('load wb')
    return load_workbook(file_name)
    #print(type(new_file))  # openpyxl.workbook.workbook.Workbook

@timer
def write_report(workbook: Workbook):
    sheet = workbook.active
    sheet['C6'] = 100

@timer
def save_report(workbook: Workbook, new_file_name):
    if os.path.isfile(new_file_name):
        os.remove(new_file_name)
    
    workbook.save(new_file_name)

wb = get_template(file_name)
write_report(wb)
save_report(wb, destination_file)


# @timer
# def write_report(destination_file):
#     if os.path.isfile(file_name):
#         new_workbook: Workbook = copy(open_workbook(file_name, formatting_info=True))
#         sheet: Sheet = new_workbook.get_sheet(0)
#         sheet.write(5, 2, 100)
#         save_workbook(new_workbook, destination_file)


# @timer
# def save_workbook(full_file_name: Workbook, destination_file):
#     if os.path.isfile(destination_file):
#         os.remove(destination_file)
#     full_file_name.save(destination_file)

# write_report(destination_file)

# if os.path.isfile(file_name):
#     print(f"file {file_name} exists")
#     wb = open_workbook(file_name)

#     sheet: Sheet = wb.sheet_by_index(0)
#     print(sheet.cell_value(rowx=5, colx=2))

# else:
#     print(f"file {file_name} does not exists")
